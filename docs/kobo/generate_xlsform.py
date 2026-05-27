#!/usr/bin/env python3
"""
Génère le fichier XLSForm XLSX pour l'enrôlement des membres FaîtièreHub.
Avec sélections en CASCADE : Région → Préfecture → Canton

Le numéro de carte et le niveau sont générés AUTOMATIQUEMENT par le système.

Usage:
    python generate_xlsform.py

Output:
    ./faitierehub_carte_membre_v2.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# COULEURS & STYLES
# =============================================================================
HEADER_FILL = PatternFill(start_color="1A6B3C", end_color="1A6B3C", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
GROUP_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
REPEAT_FILL = PatternFill(start_color="F4A800", end_color="F4A800", fill_type="solid")
NOTE_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
CALC_FILL = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
DATA_FONT = Font(name="Arial", size=9)
EVEN_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# =============================================================================
# SURVEY DATA — avec colonne choice_filter pour les cascades
# =============================================================================
# Columns: type, name, label::French, hint::French, required, constraint,
#           constraint_message::French, appearance, relevant, calculate,
#           repeat_count, choice_filter

SURVEY_HEADERS = [
    "type", "name", "label::French", "hint::French", "required",
    "constraint", "constraint_message::French", "appearance",
    "relevant", "calculate", "repeat_count", "choice_filter"
]

SURVEY_DATA = [
    # =========================================================================
    # SECTION 1 — Identité du membre
    # =========================================================================
    ["begin_group", "S1", "Section 1 — Identité du membre", "", "", "", "", "field-list", "", "", "", ""],
    ["text", "nom_complet", "Nom complet du membre", "Prénom et nom de famille", "yes", "", "", "", "", "", "", ""],
    ["date", "date_naissance", "Date de naissance", "", "yes", ". <= today()", "La date ne peut pas être dans le futur", "", "", "", "", ""],
    ["select_one sexe_list", "sexe", "Sexe", "", "yes", "", "", "minimal", "", "", "", ""],
    ["text", "telephone", "Numéro de téléphone", "Ex: +228 90 12 34 56", "yes", "regex(., '^[0-9+\\s\\-]{8,15}$')", "Numéro invalide (8 à 15 chiffres)", "", "", "", "", ""],
    ["text", "email", "Adresse email", "Optionnel", "no", "", "", "", "", "", "", ""],
    ["image", "photo_membre", "Photo portrait du membre", "Photo face, fond clair de préférence", "yes", "", "", "annotate", "", "", "", ""],
    ["select_one oui_non_list", "consentement_donnees", "Je consens au traitement de mes données personnelles par la faîtière FENOMAT", "Obligatoire pour l'enrôlement", "yes", ". = 'oui'", "Le consentement est obligatoire pour procéder à l'enrôlement", "", "", "", "", ""],
    ["end_group", "S1", "", "", "", "", "", "", "", "", "", ""],

    # =========================================================================
    # SECTION 2 — Organisation (FENOMAT pré-rempli)
    # =========================================================================
    ["begin_group", "S2", "Section 2 — Organisation", "", "", "", "", "field-list", "", "", "", ""],
    ["calculate", "code_faitiere", "", "", "", "", "", "", "", "'FENOMAT'", "", ""],
    ["note", "note_faitiere", "🏢 Faîtière : FENOMAT (Fédération Nationale des Organisations de Maraîchers du Togo)", "", "", "", "", "", "", "", "", ""],
    ["text", "nom_cooperative", "Nom de la coopérative", "Coopérative d'appartenance du membre", "yes", "", "", "", "", "", "", ""],
    ["text", "nom_union", "Nom de l'union régionale", "Si applicable", "no", "", "", "", "", "", "", ""],
    ["date", "date_adhesion", "Date d'adhésion à la coopérative", "", "yes", ". <= today()", "La date ne peut pas être dans le futur", "", "", "", "", ""],
    ["select_one statut_list", "statut_membre", "Statut du membre", "", "yes", "", "", "minimal", "", "", "", ""],
    ["note", "note_carte_auto", "ℹ️ Le numéro de carte membre et le niveau (Bronze/Argent/Or) seront générés automatiquement par le système après validation de cet enrôlement.", "", "", "", "", "", "", "", "", ""],
    ["end_group", "S2", "", "", "", "", "", "", "", "", "", ""],

    # =========================================================================
    # SECTION 3 — Localisation géographique (CASCADE: region → prefecture → canton)
    # =========================================================================
    ["begin_group", "S3", "Section 3 — Localisation géographique", "", "", "", "", "field-list", "", "", "", ""],
    ["select_one region_list", "region", "Région", "", "yes", "", "", "minimal", "", "", "", ""],
    ["select_one prefecture_list", "prefecture", "Préfecture", "", "yes", "", "", "minimal", "", "", "", "region=${region}"],
    ["select_one canton_list", "canton", "Canton", "", "no", "", "", "minimal", "", "", "", "prefecture=${prefecture}"],
    ["text", "village", "Village / Localité", "", "yes", "", "", "", "", "", "", ""],
    ["geopoint", "gps_localisation", "Position GPS", "Optionnel — améliore la cartographie", "no", "", "", "maps", "", "", "", ""],
    ["end_group", "S3", "", "", "", "", "", "", "", "", "", ""],

    # =========================================================================
    # SECTION 4 — Cotisations
    # =========================================================================
    ["begin_group", "S4", "Section 4 — Cotisations", "", "", "", "", "field-list", "", "", "", ""],
    ["integer", "nb_cotisations_12mois", "Nombre de cotisations payées dans les 12 derniers mois", "", "yes", ". >= 0", "Le nombre doit être positif ou zéro", "", "", "", "", ""],
    ["integer", "montant_derniere_cotisation", "Montant de la dernière cotisation (FCFA)", "", "no", ". >= 0", "Le montant doit être positif ou zéro", "", "", "", "", ""],
    ["date", "date_derniere_cotisation", "Date de la dernière cotisation", "", "no", ". <= today()", "La date ne peut pas être dans le futur", "", "", "", "", ""],
    ["select_one type_cotisation_list", "type_derniere_cotisation", "Type de la dernière cotisation", "", "no", "", "", "minimal", "", "", "", ""],
    ["end_group", "S4", "", "", "", "", "", "", "", "", "", ""],

    # =========================================================================
    # SECTION 5 — Parcelles agricoles (repeat)
    # =========================================================================
    ["begin_repeat", "S5", "Section 5 — Parcelles agricoles", "Ajoutez une entrée par parcelle exploitée", "", "", "", "", "", "", "", ""],
    ["select_one culture_list", "culture_principale", "Culture principale de la parcelle", "", "yes", "", "", "minimal", "", "", "", ""],
    ["decimal", "superficie_ha", "Superficie de la parcelle (hectares)", "Ex: 0.5 pour un demi-hectare", "yes", ". > 0 and . <= 500", "La superficie doit être entre 0 et 500 hectares", "", "", "", "", ""],
    ["select_one type_sol_list", "type_sol", "Type de sol", "", "yes", "", "", "minimal", "", "", "", ""],
    ["select_one irrigation_list", "irrigation", "Mode d'irrigation", "", "yes", "", "", "minimal", "", "", "", ""],
    ["select_one type_agriculture_list", "type_agriculture", "Type d'agriculture pratiquée", "", "yes", "", "", "minimal", "", "", "", ""],
    ["text", "localite_parcelle", "Localité de la parcelle", "Si différente du village de résidence", "no", "", "", "", "", "", "", ""],
    ["end_repeat", "S5", "", "", "", "", "", "", "", "", "", ""],

    # =========================================================================
    # SECTION 6 — Productions (repeat)
    # =========================================================================
    ["begin_repeat", "S6", "Section 6 — Productions de la campagne", "Ajoutez une entrée par production", "", "", "", "", "", "", "", ""],
    ["select_one culture_list", "culture_produite", "Culture produite", "", "yes", "", "", "minimal", "", "", "", ""],
    ["select_one campagne_list", "campagne_annee", "Année de la campagne", "", "yes", "", "", "minimal", "", "", "", ""],
    ["decimal", "rendement_kg", "Rendement total (kg)", "Quantité totale récoltée en kilogrammes", "yes", ". > 0", "Le rendement doit être supérieur à 0", "", "", "", "", ""],
    ["decimal", "quantite_vendue_kg", "Quantité vendue (kg)", "", "no", ". >= 0", "La quantité doit être positive ou zéro", "", "", "", "", ""],
    ["integer", "prix_vente_fcfa", "Prix de vente moyen (FCFA/kg)", "", "no", ". >= 0", "Le prix doit être positif ou zéro", "", "", "", "", ""],
    ["end_repeat", "S6", "", "", "", "", "", "", "", "", "", ""],

    # =========================================================================
    # SECTION 7 — Signature, validation & soumission
    # =========================================================================
    ["begin_group", "S7", "Section 7 — Signature et validation", "", "", "", "", "field-list", "", "", "", ""],
    ["note", "note_recap", "📝 Récapitulatif :\n• Nom : ${nom_complet}\n• Téléphone : ${telephone}\n• Coopérative : ${nom_cooperative}\n• Faîtière : FENOMAT\n• Région : ${region}\n• Préfecture : ${prefecture}\n• Village : ${village}\n\n🔑 Le numéro de carte et le niveau seront générés automatiquement après validation.", "", "", "", "", "", "", "", "", ""],
    ["image", "signature_membre", "Signature du membre", "Le membre signe sur l'écran avec le doigt", "yes", "", "", "signature", "", "", "", ""],
    ["image", "empreinte_membre", "Empreinte digitale du membre", "Alternative à la signature — le membre pose son doigt sur l'écran", "no", "", "", "draw", "", "", "", ""],
    ["select_one confirmation_list", "confirmation_soumission", "Je certifie que les informations saisies sont exactes et complètes", "", "yes", ". = 'oui'", "Vous devez certifier l'exactitude des informations pour soumettre", "", "", "", "", ""],
    ["text", "observations_agent", "Observations de l'agent collecteur", "Remarques, difficultés rencontrées, etc.", "no", "", "", "multiline", "", "", "", ""],
    ["text", "nom_agent", "Nom de l'agent collecteur", "", "yes", "", "", "", "", "", "", ""],
    ["end_group", "S7", "", "", "", "", "", "", "", "", "", ""],
]

# =============================================================================
# CHOICES DATA — avec colonnes "region" et "prefecture" pour les cascades
# =============================================================================
# Columns: list_name, name, label::French, region, prefecture
# "region" filtre les préfectures, "prefecture" filtre les cantons

CHOICES_HEADERS = ["list_name", "name", "label::French", "region", "prefecture"]

CHOICES_DATA = [
    # =========================================================================
    # Sexe
    # =========================================================================
    ["sexe_list", "homme", "Homme", "", ""],
    ["sexe_list", "femme", "Femme", "", ""],

    # =========================================================================
    # Statut membre
    # =========================================================================
    ["statut_list", "actif", "Actif", "", ""],
    ["statut_list", "inactif", "Inactif", "", ""],
    ["statut_list", "suspendu", "Suspendu", "", ""],

    # =========================================================================
    # RÉGIONS du Togo (5 régions — pas de filtre)
    # =========================================================================
    ["region_list", "maritime", "Maritime", "", ""],
    ["region_list", "plateaux", "Plateaux", "", ""],
    ["region_list", "centrale", "Centrale", "", ""],
    ["region_list", "kara", "Kara", "", ""],
    ["region_list", "savanes", "Savanes", "", ""],

    # =========================================================================
    # PRÉFECTURES du Togo — filtrées par "region"
    # =========================================================================
    # Région Maritime
    ["prefecture_list", "golfe", "Golfe (Lomé)", "maritime", ""],
    ["prefecture_list", "agoenyive", "Agoè-Nyivé", "maritime", ""],
    ["prefecture_list", "lacs", "Lacs", "maritime", ""],
    ["prefecture_list", "vo", "Vo", "maritime", ""],
    ["prefecture_list", "yoto", "Yoto", "maritime", ""],
    ["prefecture_list", "zio", "Zio", "maritime", ""],
    ["prefecture_list", "ave", "Avé", "maritime", ""],
    ["prefecture_list", "bas_mono", "Bas-Mono", "maritime", ""],
    # Région des Plateaux
    ["prefecture_list", "ogou", "Ogou", "plateaux", ""],
    ["prefecture_list", "kloto", "Kloto", "plateaux", ""],
    ["prefecture_list", "amou", "Amou", "plateaux", ""],
    ["prefecture_list", "haho", "Haho", "plateaux", ""],
    ["prefecture_list", "est_mono", "Est-Mono", "plateaux", ""],
    ["prefecture_list", "moyen_mono", "Moyen-Mono", "plateaux", ""],
    ["prefecture_list", "agou", "Agou", "plateaux", ""],
    ["prefecture_list", "danyi", "Danyi", "plateaux", ""],
    ["prefecture_list", "wawa", "Wawa", "plateaux", ""],
    ["prefecture_list", "akebou", "Akébou", "plateaux", ""],
    # Région Centrale
    ["prefecture_list", "tchaoudjo", "Tchaoudjo", "centrale", ""],
    ["prefecture_list", "tchamba", "Tchamba", "centrale", ""],
    ["prefecture_list", "sotouboua", "Sotouboua", "centrale", ""],
    ["prefecture_list", "blitta", "Blitta", "centrale", ""],
    ["prefecture_list", "mo", "Mô", "centrale", ""],
    # Région de la Kara
    ["prefecture_list", "kozah", "Kozah", "kara", ""],
    ["prefecture_list", "binah", "Binah", "kara", ""],
    ["prefecture_list", "doufelgou", "Doufelgou", "kara", ""],
    ["prefecture_list", "keran", "Kéran", "kara", ""],
    ["prefecture_list", "dankpen", "Dankpen", "kara", ""],
    ["prefecture_list", "bassar", "Bassar", "kara", ""],
    ["prefecture_list", "assoli", "Assoli", "kara", ""],
    # Région des Savanes
    ["prefecture_list", "tone", "Tône", "savanes", ""],
    ["prefecture_list", "oti", "Oti", "savanes", ""],
    ["prefecture_list", "oti_sud", "Oti-Sud", "savanes", ""],
    ["prefecture_list", "kpendjal", "Kpendjal", "savanes", ""],
    ["prefecture_list", "kpendjal_ouest", "Kpendjal-Ouest", "savanes", ""],
    ["prefecture_list", "tandjoare", "Tandjouaré", "savanes", ""],
    ["prefecture_list", "cinkasse", "Cinkassé", "savanes", ""],

    # =========================================================================
    # CANTONS — filtrés par "prefecture" (principaux cantons par préfecture)
    # =========================================================================
    # Golfe
    ["canton_list", "lome_commune", "Lomé Commune", "", "golfe"],
    ["canton_list", "togblekope", "Togblékopé", "", "golfe"],
    ["canton_list", "aflao_sagbado", "Aflao-Sagbado", "", "golfe"],
    ["canton_list", "baguida", "Baguida", "", "golfe"],
    # Agoè-Nyivé
    ["canton_list", "agoe", "Agoè", "", "agoenyive"],
    ["canton_list", "nyive", "Nyivé", "", "agoenyive"],
    ["canton_list", "legbassito", "Légbassito", "", "agoenyive"],
    # Zio
    ["canton_list", "tsevie", "Tsévié", "", "zio"],
    ["canton_list", "davie", "Davié", "", "zio"],
    ["canton_list", "gape", "Gapé", "", "zio"],
    ["canton_list", "mission_tove", "Mission-Tové", "", "zio"],
    # Vo
    ["canton_list", "vogan", "Vogan", "", "vo"],
    ["canton_list", "akoumape", "Akoumapé", "", "vo"],
    ["canton_list", "afagnan", "Afagnan", "", "vo"],
    # Yoto
    ["canton_list", "tabligbo", "Tabligbo", "", "yoto"],
    ["canton_list", "notse", "Notsé", "", "yoto"],
    # Lacs
    ["canton_list", "aneho", "Aného", "", "lacs"],
    ["canton_list", "agbodrafo", "Agbodrafo", "", "lacs"],
    # Avé
    ["canton_list", "keve", "Kévé", "", "ave"],
    ["canton_list", "assahoun", "Assahoun", "", "ave"],
    # Bas-Mono
    ["canton_list", "afagnan_bm", "Afagnan", "", "bas_mono"],
    # Ogou
    ["canton_list", "atakpame", "Atakpamé", "", "ogou"],
    ["canton_list", "okou", "Okou", "", "ogou"],
    ["canton_list", "agbandi", "Agbandi", "", "ogou"],
    # Kloto
    ["canton_list", "kpalime", "Kpalimé", "", "kloto"],
    ["canton_list", "agou_nyogbo", "Agou-Nyogbo", "", "kloto"],
    ["canton_list", "tove", "Tové", "", "kloto"],
    # Haho
    ["canton_list", "notsé_haho", "Notsé", "", "haho"],
    ["canton_list", "wahala", "Wahala", "", "haho"],
    # Amou
    ["canton_list", "amlamé", "Amlamé", "", "amou"],
    ["canton_list", "game", "Gamé", "", "amou"],
    # Wawa
    ["canton_list", "badou", "Badou", "", "wawa"],
    ["canton_list", "tomegbe", "Tomégbé", "", "wawa"],
    # Agou
    ["canton_list", "agou_gare", "Agou-Gare", "", "agou"],
    ["canton_list", "agou_nyogbo2", "Agou-Nyogbo", "", "agou"],
    # Danyi
    ["canton_list", "danyi_apeyeme", "Danyi-Apéyémé", "", "danyi"],
    ["canton_list", "danyi_elavanyo", "Danyi-Élavanyo", "", "danyi"],
    # Akébou
    ["canton_list", "kougnohou", "Kougnohou", "", "akebou"],
    # Est-Mono
    ["canton_list", "elavagnon", "Élavagnon", "", "est_mono"],
    # Moyen-Mono
    ["canton_list", "tohoun", "Tohoun", "", "moyen_mono"],
    # Tchaoudjo
    ["canton_list", "sokode", "Sokodé", "", "tchaoudjo"],
    ["canton_list", "tchalo", "Tchalo", "", "tchaoudjo"],
    ["canton_list", "komah", "Komah", "", "tchaoudjo"],
    # Tchamba
    ["canton_list", "tchamba_ville", "Tchamba", "", "tchamba"],
    ["canton_list", "alibi", "Alibi", "", "tchamba"],
    # Sotouboua
    ["canton_list", "sotouboua_ville", "Sotouboua", "", "sotouboua"],
    ["canton_list", "fazao", "Fazao", "", "sotouboua"],
    # Blitta
    ["canton_list", "blitta_ville", "Blitta", "", "blitta"],
    ["canton_list", "pagala", "Pagala", "", "blitta"],
    # Mô
    ["canton_list", "djarkpanga", "Djarkpanga", "", "mo"],
    # Kozah
    ["canton_list", "kara_ville", "Kara", "", "kozah"],
    ["canton_list", "lama_kara", "Lama-Kara", "", "kozah"],
    ["canton_list", "pya", "Pya", "", "kozah"],
    # Binah
    ["canton_list", "pagouda", "Pagouda", "", "binah"],
    ["canton_list", "kemerida", "Kémérida", "", "binah"],
    # Doufelgou
    ["canton_list", "niamtougou", "Niamtougou", "", "doufelgou"],
    ["canton_list", "siou", "Siou", "", "doufelgou"],
    # Kéran
    ["canton_list", "kanté", "Kanté", "", "keran"],
    ["canton_list", "atalote", "Atalote", "", "keran"],
    # Bassar
    ["canton_list", "bassar_ville", "Bassar", "", "bassar"],
    ["canton_list", "kabou", "Kabou", "", "bassar"],
    # Dankpen
    ["canton_list", "guerin_kouka", "Guérin-Kouka", "", "dankpen"],
    # Assoli
    ["canton_list", "bafilo", "Bafilo", "", "assoli"],
    # Tône
    ["canton_list", "dapaong", "Dapaong", "", "tone"],
    ["canton_list", "naki_est", "Naki-Est", "", "tone"],
    ["canton_list", "bombouaka", "Bombouaka", "", "tone"],
    # Oti
    ["canton_list", "mango", "Mango", "", "oti"],
    ["canton_list", "gando", "Gando", "", "oti"],
    # Oti-Sud
    ["canton_list", "sadori", "Sadori", "", "oti_sud"],
    # Kpendjal
    ["canton_list", "mandouri", "Mandouri", "", "kpendjal"],
    # Kpendjal-Ouest
    ["canton_list", "nayega", "Nayéga", "", "kpendjal_ouest"],
    # Tandjouaré
    ["canton_list", "tandjoare_ville", "Tandjouaré", "", "tandjoare"],
    ["canton_list", "nano", "Nano", "", "tandjoare"],
    # Cinkassé
    ["canton_list", "cinkasse_ville", "Cinkassé", "", "cinkasse"],

    # =========================================================================
    # Cultures maraîchères (Top 10 au Togo — FENOMAT)
    # =========================================================================
    ["culture_list", "tomate", "Tomate", "", ""],
    ["culture_list", "piment", "Piment", "", ""],
    ["culture_list", "gombo", "Gombo", "", ""],
    ["culture_list", "oignon", "Oignon", "", ""],
    ["culture_list", "laitue", "Laitue", "", ""],
    ["culture_list", "carotte", "Carotte", "", ""],
    ["culture_list", "chou", "Chou", "", ""],
    ["culture_list", "concombre", "Concombre", "", ""],
    ["culture_list", "aubergine", "Aubergine", "", ""],
    ["culture_list", "poivron", "Poivron", "", ""],
    ["culture_list", "autre", "Autre culture", "", ""],

    # =========================================================================
    # Campagnes agricoles
    # =========================================================================
    ["campagne_list", "2020", "2020", "", ""],
    ["campagne_list", "2021", "2021", "", ""],
    ["campagne_list", "2022", "2022", "", ""],
    ["campagne_list", "2023", "2023", "", ""],
    ["campagne_list", "2024", "2024", "", ""],
    ["campagne_list", "2025", "2025", "", ""],
    ["campagne_list", "2026", "2026", "", ""],

    # =========================================================================
    # Type de sol
    # =========================================================================
    ["type_sol_list", "argileux", "Argileux", "", ""],
    ["type_sol_list", "sableux", "Sableux", "", ""],
    ["type_sol_list", "limoneux", "Limoneux", "", ""],
    ["type_sol_list", "lateritique", "Latéritique", "", ""],
    ["type_sol_list", "ferralitique", "Ferralitique", "", ""],
    ["type_sol_list", "hydromorphe", "Hydromorphe (bas-fond)", "", ""],
    ["type_sol_list", "volcanique", "Volcanique", "", ""],
    ["type_sol_list", "mixte", "Mixte", "", ""],

    # =========================================================================
    # Irrigation
    # =========================================================================
    ["irrigation_list", "pluviale", "Pluviale (pluie uniquement)", "", ""],
    ["irrigation_list", "goutte_a_goutte", "Goutte-à-goutte", "", ""],
    ["irrigation_list", "aspersion", "Aspersion", "", ""],
    ["irrigation_list", "gravitaire", "Gravitaire (canal/rigole)", "", ""],
    ["irrigation_list", "pompage", "Pompage (motopompe)", "", ""],
    ["irrigation_list", "bas_fond", "Bas-fond / décrue", "", ""],
    ["irrigation_list", "aucune", "Aucune irrigation", "", ""],

    # =========================================================================
    # Type d'agriculture
    # =========================================================================
    ["type_agriculture_list", "conventionnel", "Conventionnel", "", ""],
    ["type_agriculture_list", "biologique", "Biologique", "", ""],
    ["type_agriculture_list", "agroforesterie", "Agroforesterie", "", ""],
    ["type_agriculture_list", "maraichage", "Maraîchage", "", ""],
    ["type_agriculture_list", "elevage", "Élevage", "", ""],
    ["type_agriculture_list", "pisciculture", "Pisciculture", "", ""],
    ["type_agriculture_list", "mixte", "Mixte (plusieurs types)", "", ""],

    # =========================================================================
    # Type de cotisation
    # =========================================================================
    ["type_cotisation_list", "cotisation", "Cotisation annuelle", "", ""],
    ["type_cotisation_list", "credit", "Crédit", "", ""],
    ["type_cotisation_list", "remboursement", "Remboursement", "", ""],
    ["type_cotisation_list", "amende", "Amende", "", ""],
    ["type_cotisation_list", "don", "Don", "", ""],

    # =========================================================================
    # Oui / Non
    # =========================================================================
    ["oui_non_list", "oui", "Oui", "", ""],
    ["oui_non_list", "non", "Non", "", ""],

    # Confirmation
    ["confirmation_list", "oui", "Oui, je certifie", "", ""],
    ["confirmation_list", "non", "Non", "", ""],
]

# =============================================================================
# SETTINGS DATA
# =============================================================================
SETTINGS_HEADERS = ["form_title", "form_id", "version", "default_language"]
SETTINGS_DATA = [
    ["Enrôlement Membre FENOMAT — Carte FaîtièreHub", "fenomat_enrolement_carte_v2", "2026-05-27", "French"],
]


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def apply_header_style(ws, headers):
    """Apply header styling to the first row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def get_row_fill(row_data):
    """Determine the fill color for a row based on its type."""
    if not row_data:
        return None
    row_type = row_data[0] if row_data else ""
    if row_type.startswith("begin_group") or row_type.startswith("end_group"):
        return GROUP_FILL
    if row_type.startswith("begin_repeat") or row_type.startswith("end_repeat"):
        return REPEAT_FILL
    if row_type == "note":
        return NOTE_FILL
    if row_type == "calculate":
        return CALC_FILL
    return None


def auto_width(ws, headers, data):
    """Auto-size columns based on content."""
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in data:
            if col_idx <= len(row):
                cell_len = len(str(row[col_idx - 1])) if row[col_idx - 1] else 0
                max_len = max(max_len, cell_len)
        adjusted_width = min(max_len + 2, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def write_sheet(ws, headers, data, alternate_rows=True):
    """Write headers and data to a worksheet with styling."""
    apply_header_style(ws, headers)

    for row_idx, row_data in enumerate(data, 2):
        row_fill = get_row_fill(row_data)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value else "")
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            if row_fill:
                cell.fill = row_fill
            elif alternate_rows and row_idx % 2 == 0:
                cell.fill = EVEN_FILL

    auto_width(ws, headers, data)
    ws.freeze_panes = "A2"


# =============================================================================
# MAIN — Generate the XLSX
# =============================================================================

def main():
    wb = Workbook()

    # --- Sheet 1: survey ---
    ws_survey = wb.active
    ws_survey.title = "survey"
    write_sheet(ws_survey, SURVEY_HEADERS, SURVEY_DATA)

    # --- Sheet 2: choices ---
    ws_choices = wb.create_sheet("choices")
    write_sheet(ws_choices, CHOICES_HEADERS, CHOICES_DATA)

    # --- Sheet 3: settings ---
    ws_settings = wb.create_sheet("settings")
    write_sheet(ws_settings, SETTINGS_HEADERS, SETTINGS_DATA, alternate_rows=False)

    # --- Save ---
    output_path = "faitierehub_carte_membre_v2.xlsx"
    wb.save(output_path)
    print(f"✅ XLSForm généré avec succès : {output_path}")
    print(f"   → {len(SURVEY_DATA)} lignes dans 'survey'")
    print(f"   → {len(CHOICES_DATA)} lignes dans 'choices'")
    print(f"   → Sélections en cascade : Région → Préfecture → Canton")
    print(f"   → Prêt à importer dans KoboToolbox")


if __name__ == "__main__":
    main()
